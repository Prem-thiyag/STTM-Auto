#!/usr/bin/env python3
"""
End-to-end regression check for the deterministic core of sqlx-etl-generator.

This is the driver for `.claude/skills/run-sqlx-etl-generator/`. It cannot exercise
the LLM-reasoning specialists (Schema Parser, STTM Parser's classification pass,
Dependency Builder, Mapping Resolver) -- those need an actual model in the loop.
What it CAN do, and does, is prove the entire deterministic surface still behaves
exactly as documented:

  1. scripts/parse_sttm.py actually parses a real .xlsx workbook correctly.
  2. Every artifact checked into docs/examples/generated-project/ still validates
     against its schema in schemas/.
  3. scripts/render_sqlx.py, run against the checked-in buildspecs, reproduces the
     checked-in .sqlx files BYTE FOR BYTE. This is the concrete test of ADR-001's
     "Generate is deterministic for identical inputs" claim -- not asserted, checked.
  4. scripts/gen_bootstrap.py, run against the checked-in schema IR, reproduces the
     checked-in bootstrap/ tree the same way (modulo the manifest's own timestamp).
  5. A deliberately malformed workbook is rejected with a clear error (the "fail
     loudly" contract), not silently misparsed.

Run from the skill root (the directory this file's parent's parent is):
    python scripts/smoke_test.py

Exit code 0 means every check passed. Any failure prints exactly what diverged.
"""
from __future__ import annotations

import filecmp
import json
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path

SKILL_ROOT = Path(__file__).resolve().parent.parent
EXAMPLE = SKILL_ROOT / "docs" / "examples"
REFERENCE_PROJECT = EXAMPLE / "generated-project"
PY = sys.executable

failures: list[str] = []


def run(*args: str) -> subprocess.CompletedProcess:
    return subprocess.run([PY, *args], cwd=SKILL_ROOT, capture_output=True, text=True)


def check(condition: bool, label: str, detail: str = "") -> None:
    status = "PASS" if condition else "FAIL"
    print(f"[{status}] {label}")
    if not condition:
        failures.append(f"{label}: {detail}")
        if detail:
            print(f"       {detail}")


def step1_parse_sttm() -> None:
    print("\n== 1. scripts/parse_sttm.py against the real fixture workbook ==")
    with tempfile.TemporaryDirectory() as tmp:
        out = Path(tmp) / "sttm.raw.json"
        result = run(
            "scripts/parse_sttm.py",
            str(EXAMPLE / "sttm_workbook.xlsx"),
            "--output", str(out),
        )
        check(result.returncode == 0, "parse_sttm.py exits 0 on a valid workbook", result.stderr)
        if result.returncode == 0:
            data = json.loads(out.read_text(encoding="utf-8"))
            check(len(data["rows"]) == 7, "parsed exactly 7 mapping rows", f"got {len(data.get('rows', []))}")
            multi = next(r for r in data["rows"] if r["target_column"] == "full_name")
            check(
                multi["source_column"] == ["first_name", "last_name"],
                "multi-column source parsed as a list",
                str(multi["source_column"]),
            )

    print("\n== 1b. parse_sttm.py rejects a workbook missing required headers ==")
    import openpyxl  # local import: only needed for these synthetic negative-test workbooks

    with tempfile.TemporaryDirectory() as tmp:
        bad = Path(tmp) / "bad.xlsx"
        wb = openpyxl.Workbook()
        wb.active.append(["Target Table", "Target Column", "Source Table"])
        wb.active.append(["X", "y", "z"])
        wb.save(bad)
        result = run("scripts/parse_sttm.py", str(bad))
        check(result.returncode == 2, "exits 2 on missing required headers", result.stderr)
        check("header contract violation" in result.stderr, "error names the header contract", result.stderr)

    print("\n== 1c. parse_sttm.py accepts a legitimate sourceless (GENERATED) row and flags duplicates ==")
    headers = ["Target Table", "Target Column", "Source Table", "Source Column", "Transformation"]
    with tempfile.TemporaryDirectory() as tmp:
        wb_path = Path(tmp) / "generated_column.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(["T", "surrogate_key", None, None, "Surrogate Key"])  # both blank: legitimate
        ws.append(["T", "id", "S", "id", "Direct"])
        ws.append(["T", "dup_col", "S", "a", "Direct"])
        ws.append(["T", "dup_col", "S", "b", "Direct"])  # same (table, column) as the row above
        wb.save(wb_path)
        out = Path(tmp) / "out.json"
        result = run("scripts/parse_sttm.py", str(wb_path), "--output", str(out))
        check(result.returncode == 0, "exits 0 on a both-blank source row", result.stderr)
        if result.returncode == 0:
            rows = {r["target_column"]: r for r in json.loads(out.read_text(encoding="utf-8"))["rows"]}
            check(
                rows["surrogate_key"]["source_table"] is None and rows["surrogate_key"]["source_column"] is None,
                "sourceless row parses with source_table/source_column both null",
                str(rows["surrogate_key"]),
            )
            check(rows["surrogate_key"]["duplicate_mapping"] is False, "sourceless row is not flagged duplicate")
            check(rows["id"]["duplicate_mapping"] is False, "unique target_column is not flagged duplicate")
            dup_rows = [r for r in json.loads(out.read_text(encoding="utf-8"))["rows"] if r["target_column"] == "dup_col"]
            check(
                len(dup_rows) == 2 and all(r["duplicate_mapping"] for r in dup_rows),
                "both rows sharing (target_table, target_column) are flagged duplicate_mapping=true",
                str(dup_rows),
            )

    print("\n== 1d. parse_sttm.py rejects a row with exactly one of Source Table/Source Column blank ==")
    with tempfile.TemporaryDirectory() as tmp:
        wb_path = Path(tmp) / "inconsistent_source.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        ws.append(["T", "bad_col", "S", None, "Direct"])  # source_table without source_column
        wb.save(wb_path)
        result = run("scripts/parse_sttm.py", str(wb_path))
        check(result.returncode == 3, "exits 3 on an inconsistent partial-source row", result.stderr)
        check("inconsistent source" in result.stderr, "error names the inconsistency", result.stderr)


def step2_validate_reference_artifacts() -> None:
    print("\n== 2. every checked-in reference artifact validates against its schema ==")
    pairs = [
        (REFERENCE_PROJECT / "metadata/schema/source_schema.json", "schemas/source_schema.schema.json"),
        (REFERENCE_PROJECT / "metadata/schema/target_schema.json", "schemas/target_schema.schema.json"),
        (REFERENCE_PROJECT / "metadata/mapping/sttm.json", "schemas/sttm.schema.json"),
        (REFERENCE_PROJECT / "metadata/dependency/dependency_graph.json", "schemas/dependency_graph.schema.json"),
        (REFERENCE_PROJECT / "metadata/execution/execution_plan.json", "schemas/execution_plan.schema.json"),
        (REFERENCE_PROJECT / "metadata/execution/execution_log.json", "schemas/execution_log.schema.json"),
        (REFERENCE_PROJECT / "metadata/review/review_spec.json", "schemas/review_spec.schema.json"),
        (REFERENCE_PROJECT / "metadata/review/review_report.json", "schemas/review_report.schema.json"),
        (REFERENCE_PROJECT / "metadata/cleanup/cleanup_manifest.json", "schemas/cleanup_manifest.schema.json"),
        (REFERENCE_PROJECT / "metadata/cleanup/cleanup_log.json", "schemas/cleanup_log.schema.json"),
        (REFERENCE_PROJECT / "metadata/manifest.json", "schemas/manifest.schema.json"),
        (REFERENCE_PROJECT / "bootstrap/manifest.json", "schemas/bootstrap_manifest.schema.json"),
    ]
    for instance, schema in pairs:
        result = run("scripts/validate_schema.py", str(instance), schema)
        check(result.returncode == 0, f"{instance.relative_to(REFERENCE_PROJECT)} valid", result.stderr)

    result = run(
        "scripts/validate_schema.py",
        str(REFERENCE_PROJECT / "metadata/build"),
        "schemas/buildspec.schema.json",
        "--glob", "*.buildspec.json",
    )
    check(result.returncode == 0, "both buildspecs valid", result.stderr)


def step3_render_is_reproducible() -> None:
    print("\n== 3. render_sqlx.py reproduces the checked-in definitions/ byte-for-byte ==")
    with tempfile.TemporaryDirectory() as tmp:
        out_dir = Path(tmp) / "definitions"
        result = run(
            "scripts/render_sqlx.py",
            str(REFERENCE_PROJECT / "metadata" / "build"),
            "--templates-dir", "templates/sqlx",
            "--output-dir", str(out_dir),
            "--schema", "schemas/buildspec.schema.json",
            "--intermediate-database", "intermediate_db",
            "--intermediate-schema", "staging",
        )
        check(result.returncode == 0, "render_sqlx.py exits 0", result.stderr)
        if result.returncode != 0:
            return
        reference_dir = REFERENCE_PROJECT / "definitions"
        cmp = filecmp.dircmp(str(reference_dir), str(out_dir))
        identical = not cmp.left_only and not cmp.right_only and not cmp.diff_files
        for table_dir in reference_dir.iterdir():
            for f in table_dir.iterdir():
                rel = f.relative_to(reference_dir)
                a, b = reference_dir / rel, out_dir / rel
                same = a.exists() and b.exists() and a.read_bytes() == b.read_bytes()
                if not same:
                    identical = False
        check(identical, "re-rendered .sqlx files match docs/examples/generated-project/definitions/ exactly")


def step4_bootstrap_is_reproducible() -> None:
    print("\n== 4. gen_bootstrap.py reproduces the checked-in bootstrap/ (modulo timestamps) ==")
    with tempfile.TemporaryDirectory() as tmp:
        out_dir = Path(tmp) / "bootstrap"
        result = run(
            "scripts/gen_bootstrap.py",
            "--source-schema", str(REFERENCE_PROJECT / "metadata/schema/source_schema.json"),
            "--target-schema", str(REFERENCE_PROJECT / "metadata/schema/target_schema.json"),
            "--udf-doc", str(EXAMPLE / "udf.md"),
            "--templates-dir", "templates/bootstrap",
            "--output-dir", str(out_dir),
            "--intermediate-database", "intermediate_db",
            "--intermediate-schema", "staging",
            "--project-name", "retail-analytics-etl",
        )
        check(result.returncode == 0, "gen_bootstrap.py exits 0", result.stderr)
        if result.returncode != 0:
            return

        reference_dir = REFERENCE_PROJECT / "bootstrap"
        mismatches = []
        for f in reference_dir.rglob("*"):
            if f.is_dir():
                continue
            rel = f.relative_to(reference_dir)
            other = out_dir / rel
            if not other.exists():
                mismatches.append(f"missing: {rel}")
                continue
            if rel.name == "manifest.json":
                a = json.loads(f.read_text(encoding="utf-8"))
                b = json.loads(other.read_text(encoding="utf-8"))
                a.pop("generated_at", None)
                b.pop("generated_at", None)
                if a != b:
                    mismatches.append(f"content differs (ignoring generated_at): {rel}")
            elif f.read_bytes() != other.read_bytes():
                mismatches.append(f"content differs: {rel}")
        check(not mismatches, "bootstrap/ matches reference exactly", "; ".join(mismatches))


def step5_generated_column_renders() -> None:
    print("\n== 5. render_sqlx.py renders a GENERATED (ROW_NUMBER) column correctly ==")
    with tempfile.TemporaryDirectory() as tmp:
        build_dir = Path(tmp) / "build"
        build_dir.mkdir()
        buildspec = {
            "target_table": "T",
            "target_database": "target_db",
            "target_schema": "warehouse",
            "load_strategy": "full_load",
            "source_tables": [{"database": "source_db", "table": "S"}],
            "staging_table": "stg_t",
            "columns": [
                {
                    "target_column": "surrogate_key", "type": "BIGINT", "source": None,
                    "transformation": "GENERATED", "expression": None, "udf": None,
                    "generator": {"type": "ROW_NUMBER", "order_by": "id"},
                },
                {
                    "target_column": "id", "type": "INT",
                    "source": {"table": "S", "column": "id"},
                    "transformation": "DIRECT", "expression": None, "udf": None, "generator": None,
                },
            ],
            "joins": [], "filters": [], "grain": "one row per S", "mapping_refs": ["M001", "M002"],
        }
        (build_dir / "T.buildspec.json").write_text(json.dumps(buildspec), encoding="utf-8")
        out_dir = Path(tmp) / "definitions"
        result = run(
            "scripts/render_sqlx.py", str(build_dir),
            "--templates-dir", "templates/sqlx",
            "--output-dir", str(out_dir),
            "--schema", "schemas/buildspec.schema.json",
            "--intermediate-database", "intermediate_db",
            "--intermediate-schema", "staging",
        )
        check(result.returncode == 0, "render_sqlx.py exits 0 for a GENERATED column", result.stderr)
        if result.returncode != 0:
            return
        read_sql = (out_dir / "T" / "read.sqlx").read_text(encoding="utf-8")
        check(
            "ROW_NUMBER() OVER (ORDER BY fdw_source_db.S.id) AS surrogate_key" in read_sql,
            "GENERATED ROW_NUMBER column renders the expected qualified expression",
            read_sql,
        )
        check('"version": "1.0"' in read_sql, "config block carries the version field")


def main() -> int:
    print(f"sqlx-etl-generator smoke test - skill root: {SKILL_ROOT}")
    step1_parse_sttm()
    step2_validate_reference_artifacts()
    step3_render_is_reproducible()
    step4_bootstrap_is_reproducible()
    step5_generated_column_renders()

    print(f"\n{'=' * 60}")
    if failures:
        print(f"{len(failures)} FAILURE(S):")
        for f in failures:
            print(f"  - {f}")
        return 1
    print("ALL CHECKS PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
