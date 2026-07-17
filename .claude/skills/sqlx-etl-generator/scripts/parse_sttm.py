#!/usr/bin/env python3
"""
Deterministic STTM workbook -> normalized JSON preprocessor.

Owned by the STTM Parser specialist (see specialists/sttm-parser.md). This script does
ZERO semantic interpretation: it does not classify transformation types, does not resolve
UDF references, does not guess at ambiguous cells. It only flattens the workbook's rows
into a normalized, machine-checkable structure and fails loudly on anything that doesn't
match the expected header contract.

A row's Source Table / Source Column cells may legitimately both be blank -- that's a
"generated column" mapping (a surrogate key, a UUID, a fixed default: see
schemas/sttm.schema.json's GENERATED/DEFAULT/SEQUENCE/UUID transformation kinds), not an
error. What's still a hard error is exactly ONE of the pair being blank -- that's
inconsistent, not a legitimate sourceless mapping, and reading it as one would silently
misrepresent the row. See "row-level validation" below.

The STTM Parser specialist's LLM step reads this script's output (never the workbook
directly) and classifies each row's transformation into the enum required by
schemas/sttm.schema.json (DIRECT | EXPRESSION | UDF | CONSTANT | LOOKUP | NEEDS_REVIEW),
producing the final metadata/mapping/sttm.json.

This keeps the least-structured of the five inputs (a spreadsheet) out of the LLM's
context entirely except as compact, pre-validated JSON -- the mechanism behind ADR-001's
"low token usage" and "deterministic behavior" goals for STTM ingestion.

Usage:
    python parse_sttm.py <workbook.xlsx> [--sheet SHEET_NAME] [--output OUT.json]

Exit codes:
    0  success, JSON written to --output or stdout
    1  workbook or sheet not found
    2  header contract violation (missing required column headers)
    3  row-level validation failure (Target Table/Target Column blank, or exactly
       one of Source Table/Source Column blank -- see module docstring)
"""
from __future__ import annotations

import argparse
import datetime
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print(
        "ERROR: openpyxl is required. Install with: pip install openpyxl",
        file=sys.stderr,
    )
    sys.exit(1)

# The STTM workbook's header row must contain these columns (case-insensitive,
# whitespace/punctuation-insensitive). This is the strict contract referenced in
# ADR-001 Risk table: "STTM workbook layout changes between versions" -> the script
# fails loudly here rather than silently mis-parsing a reshuffled or renamed column.
REQUIRED_HEADERS = {
    "target_table": ["target table", "target_table", "tgt table", "tgt_table"],
    "target_column": ["target column", "target_column", "tgt column", "tgt_column"],
    "source_table": ["source table", "source_table", "src table", "src_table"],
    "source_column": ["source column", "source_column", "src column", "src_column"],
    "transformation_note": [
        "transformation",
        "transformation logic",
        "transformation_logic",
        "transformation note",
        "mapping logic",
        "business logic",
    ],
}

OPTIONAL_HEADERS = {
    "udf_reference": ["udf", "udf reference", "udf_reference", "function", "function reference"],
    "notes": ["notes", "note", "comments", "comment"],
}


class HeaderContractError(ValueError):
    """The workbook's header row is missing a required column (exit code 2)."""


class RowValidationError(ValueError):
    """One or more data rows fail row-level validation (exit code 3) -- distinct
    from a header contract violation so `main()` can report the documented exit
    code for each rather than collapsing both into one generic failure."""


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def build_header_index(header_row: list) -> dict[str, int]:
    normalized = [normalize_header(v) for v in header_row]
    index: dict[str, int] = {}

    def find(aliases: list[str]) -> int | None:
        for alias in aliases:
            if alias in normalized:
                return normalized.index(alias)
        return None

    missing = []
    for field, aliases in REQUIRED_HEADERS.items():
        pos = find(aliases)
        if pos is None:
            missing.append(f"{field} (expected one of: {', '.join(aliases)})")
        else:
            index[field] = pos

    if missing:
        raise HeaderContractError(
            "STTM workbook header contract violation. Missing required column(s):\n  - "
            + "\n  - ".join(missing)
            + f"\nFound headers: {header_row}"
        )

    for field, aliases in OPTIONAL_HEADERS.items():
        pos = find(aliases)
        if pos is not None:
            index[field] = pos

    return index


def cell_str(row: tuple, index: dict[str, int], field: str) -> str | None:
    pos = index.get(field)
    if pos is None or pos >= len(row):
        return None
    value = row[pos]
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def split_multi(value: str) -> list[str] | str:
    """Source Column cells may contain a comma- or plus-separated list for
    multi-column transformations (e.g. concatenation). A single column stays a
    plain string; more than one becomes a list, matching sttm.schema.json's
    'source_column' union type."""
    parts = [p.strip() for p in value.replace("+", ",").split(",") if p.strip()]
    return parts if len(parts) > 1 else parts[0]


def parse_workbook(path: Path, sheet_name: str | None) -> dict:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found. Available sheets: {workbook.sheetnames}"
            )
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.worksheets[0]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ValueError("Workbook sheet is empty; no header row found.")

    index = build_header_index(list(header_row))

    rows = []
    row_errors = []
    for excel_row_number, row in enumerate(rows_iter, start=2):
        if row is None or all(v is None for v in row):
            continue  # skip fully blank rows

        target_table = cell_str(row, index, "target_table")
        target_column = cell_str(row, index, "target_column")
        source_table = cell_str(row, index, "source_table")
        source_column_raw = cell_str(row, index, "source_column")

        missing_required = [
            name
            for name, value in [("target_table", target_table), ("target_column", target_column)]
            if value is None
        ]
        if missing_required:
            row_errors.append(f"row {excel_row_number}: missing {', '.join(missing_required)}")
            continue

        # Source Table and Source Column must be both-blank (a legitimate
        # generated/sourceless mapping) or both-populated -- exactly one blank
        # is an inconsistent, malformed row, not a sourceless mapping.
        has_source_table = source_table is not None
        has_source_column = source_column_raw is not None
        if has_source_table != has_source_column:
            row_errors.append(
                f"row {excel_row_number}: inconsistent source -- "
                f"{'source_table without source_column' if has_source_table else 'source_column without source_table'} "
                f"(a generated/sourceless mapping needs BOTH blank, not just one)"
            )
            continue

        rows.append(
            {
                "row_number": excel_row_number,
                "target_table": target_table,
                "target_column": target_column,
                "source_table": source_table,
                "source_column": split_multi(source_column_raw) if source_column_raw is not None else None,
                "transformation_note": cell_str(row, index, "transformation_note") or "",
                "udf_reference": cell_str(row, index, "udf_reference"),
                "notes": cell_str(row, index, "notes"),
            }
        )

    if row_errors:
        raise RowValidationError(
            "STTM workbook row validation failed:\n  - " + "\n  - ".join(row_errors)
        )

    if not rows:
        raise ValueError("STTM workbook contains a valid header but zero data rows.")

    # Flag (never drop or merge) rows that share a (target_table, target_column)
    # pair -- a mechanical, deterministic signal the STTM Parser specialist's
    # classification step uses to route every row in the group to NEEDS_REVIEW
    # rather than silently picking one (see specialists/sttm-parser.md
    # "Duplicate mappings").
    target_key_counts: dict[tuple[str, str], int] = {}
    for r in rows:
        key = (r["target_table"], r["target_column"])
        target_key_counts[key] = target_key_counts.get(key, 0) + 1
    for r in rows:
        r["duplicate_mapping"] = target_key_counts[(r["target_table"], r["target_column"])] > 1

    return {
        "generated_at": datetime.datetime.now(datetime.timezone.utc)
        .isoformat()
        .replace("+00:00", "Z"),
        "source_file": str(path),
        "sheet": sheet.title,
        "rows": rows,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path, help="Path to the STTM .xlsx workbook")
    parser.add_argument("--sheet", default=None, help="Worksheet name (default: first sheet)")
    parser.add_argument(
        "--output", type=Path, default=None, help="Output JSON path (default: stdout)"
    )
    args = parser.parse_args()

    if not args.workbook.exists():
        print(f"ERROR: workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        result = parse_workbook(args.workbook, args.sheet)
    except RowValidationError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 3
    except ValueError as exc:
        # HeaderContractError and every other ValueError this module raises
        # (missing sheet, empty sheet, zero data rows) share exit code 2 --
        # only row-level validation gets its own code, per the module
        # docstring's Exit codes table.
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    payload = json.dumps(result, indent=2, ensure_ascii=False)
    if args.output:
        args.output.write_text(payload, encoding="utf-8")
        print(f"Wrote {len(result['rows'])} normalized row(s) to {args.output}", file=sys.stderr)
    else:
        print(payload)

    return 0


if __name__ == "__main__":
    sys.exit(main())
